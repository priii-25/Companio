import tensorflow as tf
import numpy as np
import cv2
import os
import sys
import json
from mtcnn import MTCNN
import openpyxl
from ultralytics import YOLO
import time

# TensorFlow session setup
config = tf.compat.v1.ConfigProto(log_device_placement=False)
session = tf.compat.v1.Session(config=config)
sess = session

# Initialize MTCNN and YOLO
detector = MTCNN()
yolo_model = YOLO("yolov8n.pt")

# Constants
minsize = 20
threshold = [0.6, 0.7, 0.7]
factor = 0.709
margin = 44
input_image_size = 160

# Load or create Excel database
loc = os.path.join("data.xlsx")
if not os.path.exists(loc):
    print("No database file found. Creating new data.xlsx.")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["ID", "Name"])
    workbook.save(loc)
else:
    workbook = openpyxl.load_workbook(loc)
    sheet = workbook.active

def custom_load_model(model):
    model_exp = os.path.expanduser(model)
    print('Model filename: %s' % model_exp)
    with tf.io.gfile.GFile(model_exp, 'rb') as f:
        graph_def = tf.compat.v1.GraphDef()
        graph_def.ParseFromString(f.read())
        tf.compat.v1.import_graph_def(graph_def, name='')

model_path = os.path.join("20170512-110547.pb")
custom_load_model(model_path)
images_placeholder = tf.compat.v1.get_default_graph().get_tensor_by_name("input:0")
embeddings = tf.compat.v1.get_default_graph().get_tensor_by_name("embeddings:0")
phase_train_placeholder = tf.compat.v1.get_default_graph().get_tensor_by_name("phase_train:0")
embedding_size = embeddings.get_shape()[1]

def prewhiten(x):
    mean = np.mean(x)
    std = np.std(x)
    std_adj = np.maximum(std, 1.0 / np.sqrt(x.size))
    y = np.multiply(np.subtract(x, mean), 1 / std_adj)
    return y

def getFace(img):
    faces = []
    try:
        result = detector.detect_faces(img)
        if not result:
            return faces
        bounding_box = result[0]['box']
        keypoints = result[0]['keypoints']
        cropped = img[bounding_box[1]:bounding_box[1] + bounding_box[3], bounding_box[0]:bounding_box[0] + bounding_box[2]]
        if cropped.size == 0:
            return faces
        rearranged = cv2.resize(cropped, (input_image_size, input_image_size), interpolation=cv2.INTER_CUBIC)
        prewhitened = prewhiten(rearranged)
        faces.append({'face': rearranged, 'embedding': getEmbedding(prewhitened), 'box': bounding_box})
    except Exception as e:
        print(f"Error in getFace: {e}")
        sys.stdout.flush()
    return faces

def getEmbedding(resized):
    try:
        reshaped = resized.reshape(-1, input_image_size, input_image_size, 3)
        feed_dict = {images_placeholder: reshaped, phase_train_placeholder: False}
        with sess.as_default():
            embedding = sess.run(embeddings, feed_dict=feed_dict)
        return embedding
    except Exception as e:
        print(f"Error in getEmbedding: {e}")
        sys.stdout.flush()
        return None

# Load database images
database_faces = []
nrows = sheet.max_row
for j in range(1, nrows):
    img_path = os.path.join("database", f"Person{j}.jpg")
    img = cv2.imread(img_path)
    if img is None:
        continue
    faces = getFace(img)
    if faces:
        name = sheet.cell(j + 1, 2).value
        database_faces.append({'embedding': faces[0]['embedding'], 'class': j, 'name': name})

print("Total number of people in database = " + str(len(database_faces)))

def segment_image(image):
    sub_images = []
    face_positions = []
    try:
        results = yolo_model(image)
        for result in results:
            boxes = result.boxes.xyxy.cpu().numpy()
            classes = result.boxes.cls.cpu().numpy()
            for box, cls in zip(boxes, classes):
                if int(cls) == 0:  # Person class
                    x1, y1, x2, y2 = map(int, box[:4])
                    face_height = int((y2 - y1) * 0.2)
                    face_y1 = y1
                    face_y2 = y1 + face_height
                    face_x1 = x1
                    face_x2 = x2
                    face_center_x = (face_x1 + face_x2) // 2
                    face_positions.append((face_center_x, face_x1, face_x2, face_y1, face_y2))
    except Exception as e:
        print(f"Error in segment_image: {e}")
        sys.stdout.flush()
        return [image]
    if not face_positions:
        return [image]
    face_positions.sort(key=lambda x: x[0])
    num_faces = len(face_positions)
    img_height, img_width = image.shape[:2]
    if num_faces == 1:
        sub_images.append(image)
    else:
        split_points = []
        for i in range(num_faces - 1):
            split_x = (face_positions[i][0] + face_positions[i + 1][0]) // 2
            split_points.append(split_x)
        prev_x = 0
        for i in range(num_faces):
            if i < len(split_points):
                next_x = split_points[i]
            else:
                next_x = img_width
            sub_img = image[:, prev_x:next_x]
            if sub_img.size > 0:
                sub_images.append(sub_img)
            prev_x = next_x
    return sub_images

def recognize_face(sub_image, sub_img_index):
    recognized_name = "Unknown"
    min_distance = float('inf')
    threshold = 0.9
    try:
        live_faces = getFace(sub_image.copy())
        if live_faces and live_faces[0]['embedding'] is not None:
            live_embedding = live_faces[0]['embedding']
            for db_face in database_faces:
                if db_face['embedding'] is None:
                    continue
                distance = np.sqrt(np.sum(np.square(np.subtract(live_embedding, db_face['embedding']))))
                if distance < min_distance:
                    min_distance = distance
                    if distance <= threshold:
                        recognized_name = db_face['name']
            if recognized_name == "Unknown":
                print(json.dumps({"status": "new_person_detected", "sub_image_index": sub_img_index}))
                sys.stdout.flush()
                # Wait for name from server (via stdin)
                start_time = time.time()
                timeout = 30  # Wait 30 seconds for name
                new_name = None
                while time.time() - start_time < timeout:
                    try:
                        new_name = input().strip()
                        if new_name:
                            break
                    except EOFError:
                        time.sleep(1)
                if new_name:
                    recognized_name = new_name
                    new_person_id = len(database_faces) + 1
                    new_img_path = os.path.join("database", f"Person{new_person_id}.jpg")
                    cv2.imwrite(new_img_path, live_faces[0]['face'])
                    database_faces.append({'embedding': live_embedding, 'class': new_person_id, 'name': new_name})
                    wb = openpyxl.load_workbook(loc)
                    ws = wb.active
                    ws.append([new_person_id, new_name])
                    wb.save(loc)
        return recognized_name
    except Exception as e:
        print(f"Error in recognize_face: {e}")
        sys.stdout.flush()
        return "Unknown"

if __name__ == "__main__":
    val = input("Enter your value from Keyboard: ")
    print(val)
    if val == '4':
        try:
            camera = cv2.VideoCapture(0)
            if not camera.isOpened():
                print("Warning: Could not open webcam.")
                sys.exit(0)
            while True:
                ret, frame = camera.read()
                if not ret:
                    print("Warning: Could not read frame.")
                    sys.exit(0)
                cv2.putText(frame, "Press 'c' to capture", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                cv2.imshow('Webcam Capture', frame)
                key = cv2.waitKey(1) & 0xFF
                if key == ord('c'):
                    cv2.imwrite("captured_frame.jpg", frame)
                    print(json.dumps({"status": "frame_captured", "path": "captured_frame.jpg"}))
                    sys.stdout.flush()
                    break
                elif key == ord('q'):
                    sys.exit(0)
            camera.release()
            cv2.destroyAllWindows()

            renamed_file_path = input().strip()
            captured_image = cv2.imread(renamed_file_path)
            if captured_image is None:
                print("Warning: Could not load captured image.")
                sys.exit(0)

            sub_images = segment_image(captured_image)
            print(f"Segmented into {len(sub_images)} sub-images")

            face_results = []
            for i, sub_img in enumerate(sub_images):
                recognized_name = recognize_face(sub_img, i)
                face_results.append({
                    "recognized": recognized_name != "Unknown",
                    "name": recognized_name if recognized_name != "Unknown" else None
                })

            result = {
                "status": "success",
                "faces": face_results
            }
            print(json.dumps(result))
            sys.stdout.flush()
            sys.exit(0)
        except Exception as e:
            print(f"Error in main execution: {e}")
            sys.stdout.flush()
            sys.exit(1)